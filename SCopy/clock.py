import re
import os

import openpyxl
from collections import defaultdict

from lxml import etree as ET


__all__ = ["Class_Clock_config"]
# def clockxl_load(self, xlpath)
# def indent(self, elem, level=0)
# def parsing_pll(self, xmlpath, ConfigItem_dict)

class Class_Clock_config:
    def __init__(self):
        super().__init__()

        return

    def __del__(self):
        return

    def clockxl_load(self, xlpath):
        excel_document = openpyxl.load_workbook(xlpath)       
        ActivateSheet = excel_document.get_sheet_by_name("RESULTS")
        # DEBUG
        # print(excel_document.get_sheet_names())
        # print(excel_document.get_sheet_by_name("RESULTS"))
        # print(ActivateSheet.max_row)
        # print(ActivateSheet.max_column)
        ConfigItem_list = list()          
        ConfigItem_dict = defaultdict(dict)
        for j in range(1,ActivateSheet.max_column):
            for i in range(1,ActivateSheet.max_row):
                SheetItem = str(ActivateSheet.cell(row=i, column=j).value).strip()
                Sheetvalue = ActivateSheet.cell(row=i, column=j+1).value
                if re.match('Mcu*', SheetItem) is not None:
                    if isinstance(Sheetvalue, int):
                            #change from normal to scientific
                            if re.search('Frequency', SheetItem) is not None:
                                temp_float = float(Sheetvalue)
                                temp_float = re.sub('\+0','',str("{:.1E}".format(temp_float)))
                                ConfigItem_dict[SheetItem].update({'value':temp_float})
                            else:              
                                ConfigItem_dict[SheetItem].update({'value':str(Sheetvalue)})
                            ConfigItem_dict[SheetItem].update({'row':i})
                            ConfigItem_dict[SheetItem].update({'column':j})
                            ConfigItem_list.append(SheetItem)
                    else:
                        #Special case
                        if re.fullmatch(Sheetvalue ,'NOMAL') is True:
                            ConfigItem_dict[SheetItem].update({'value':'NORMAL_MODE'})
                            ConfigItem_dict[SheetItem].update({'row':i})
                            ConfigItem_dict[SheetItem].update({'column':j})
                            ConfigItem_list.append(SheetItem)
                        else:
                            #skip NA
                            pass
                        
                else:
                    pass
        # DEBUG
        # for ConfigItem in ConfigItem_dict:    
        #     print(ConfigItem,'value:', ConfigItem_dict[ConfigItem]['value'],'column:', ConfigItem_dict[ConfigItem]['column'], 'row:',ConfigItem_dict[ConfigItem]['row'])
        return ConfigItem_dict

    def indent(self, elem, level=0):
        i = "\n" + (level - 1) * "  "
        j = "\n" + (level - 1) * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for subelem in elem:
                self.indent(subelem, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = j
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = j
        return elem

    def parsing_pll(self, xmlpath, ConfigItem_dict):
        tag_ctr06dataxsd = r'{http://www.tresos.de/_projects/DataModel2/06/data.xsd}ctr'
        tag_var06dataxsd = r'{http://www.tresos.de/_projects/DataModel2/06/data.xsd}var'
        tag_a08attributexsd = r'{http://www.tresos.de/_projects/DataModel2/08/attribute.xsd}a'
        
        tree = ET.parse(xmlpath)
        TaskingRoot = tree.getroot()
        self.indent(TaskingRoot)
        xmlpath_temp = os.path.split(xmlpath)
        xmlpath_temp = os.path.join(xmlpath_temp[0], "Mcu.bak")
        xmlpath_temp = xmlpath_temp.replace('\\', '/')
        tree.write(xmlpath_temp)
        tree = ET.parse(xmlpath)
        TaskingRoot = tree.getroot()
        for Level_1st in TaskingRoot.iter(tag_ctr06dataxsd):
            if 'name' in Level_1st.attrib:
                if re.fullmatch(Level_1st.get('name'), 'McuClockReferencePoint') is not None:
                    temp_name = Level_1st.get('name')
                    temp_type = Level_1st.get('type')
                    temp_value = Level_1st.get('value')
                    # print("name:", temp_name,"type:", temp_type,"value:", temp_value)              
                    for Level_2nd in Level_1st.iter(tag_var06dataxsd ):
                        idx_num = Level_1st.getchildren().index(Level_2nd)
                        # print("index Num:", idx_num)
                        for ConfigItem in ConfigItem_dict: 
                            if re.fullmatch(Level_2nd.get('name'), ConfigItem,re.IGNORECASE) is not None:
                                temp_name = Level_2nd.get('name')
                                temp_type = Level_2nd.get('type')
                                temp_value = Level_2nd.get('value')
                                # print("name:", temp_name,"type:", temp_type,"value:", temp_value)
                                # print("name:", ConfigItem, "value:", ConfigItem_dict[ConfigItem]['value'])
                                if (ConfigItem_dict[ConfigItem]['value'] == temp_value) is False:
                                    # print("change","name:", temp_name,"type:", temp_type,"value:", temp_value,"to", ConfigItem_dict[ConfigItem]['value'])                        
                                    for Level_3rd in Level_2nd.iter(tag_a08attributexsd):
                                        # print("name:", temp_name,"type:", temp_type,"value:", temp_value)
                                        # print("Deleted", Level_3rd.get('name'), Level_3rd.get('value'))
                                        Level_3rd.getparent().remove(Level_3rd)                                    
                                    Level_2nd.getparent().remove(Level_2nd)
                                    Level_2nd_element = ET.SubElement(Level_1st, tag_var06dataxsd, name = temp_name, type = temp_type, value = ConfigItem_dict[ConfigItem]['value'])                      
                                    Level_1st.insert(idx_num, Level_2nd_element)                                
                                ConfigItem_dict.pop(ConfigItem)
                                break
                    break
        self.indent(TaskingRoot)                   
        tree.write(xmlpath)
        return


